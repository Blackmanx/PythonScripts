import openpyxl
from tkinter import Tk, filedialog
from tabulate import tabulate
import sys

# Check if an argument is provided for the file path
if len(sys.argv) > 1:
	file_path = sys.argv[1]
else:
	# Ask the user to select the Excel file
	root = Tk()
	root.withdraw()
	file_path = filedialog.askopenfilename(title="Select Excel file")

# Load the workbook
workbook = openpyxl.load_workbook(file_path)

# Get the active sheet
sheet = workbook.active

# Retrieve the field names from the first row
field_names = []
for cell in sheet[1]:
	field_names.append(cell.value)

# Offer the user to select an option
print("Please select an option:")
print("1. Retrieve data for a specific field")
print("2. Show all rows that contain a value")
print("3. Show all values")
print("4. Exit")

# Get the user's selection
selected_option = int(input("Enter the number of the selected option: "))

if selected_option == 1:
	# Offer the user to select a field from the menu
	print("Please select a field:")
	for i, field in enumerate(field_names):
		print(f"{i+1}. {field}")

	# Get the user's selection
	selected_index = int(input("Enter the number of the selected field: "))
	selected_field = field_names[selected_index - 1]

	# Find the column number for the selected field
	field_column = None
	for cell in sheet[1]:
		if cell.value == selected_field:
			field_column = cell.column_letter
			break

	# Retrieve the data from the selected column for all rows
	data = []
	for row in sheet.iter_rows(min_row=2):
		cell_value = row[sheet[field_column+'1'].col_idx - 1].value
		data.append(cell_value)

	# Print the retrieved data
	print(f"Retrieved data for {selected_field}:", data)

	# Ask the user if they want to store the output in a text file
	store_output = input("Do you want to store the output in a text file? (yes/no): ")

	if store_output.lower() == "yes" or store_output.lower() == "y":
		# Ask for the name of the text file
		file_name = input("Enter the name of the text file: ")

		# Create the text file if it doesn't exist
		with open(file_name, "w") as file:
			file.write(f"Retrieved data for {selected_field}: {data}")

elif selected_option == 2:
	# Ask the user for the value to search
	search_value = input("Enter the value to search: ")

	# Retrieve all rows that contain the search value
	rows_with_value = []
	for row in sheet.iter_rows(min_row=2):
		for cell in row:
			if str(cell.value) == search_value:
				rows_with_value.append([cell.value for cell in row])
				break

	# Ask the user if they want the output in table format or not
	output_format = input("Do you want the output in table format? (yes/no): ")

	if output_format.lower() == "yes" or output_format.lower() == "y":
		# Print the rows that contain the search value in a table format
		table = []
		for row in rows_with_value:
			table.append(row)
		print(tabulate(table, headers=field_names))

	else:
		# Print the rows that contain the search value without table format
		print(f"Rows that contain the value '{search_value}':")
		for row in rows_with_value:
			print(row)

	# Ask the user if they want to store the output in a text file
	store_output = input("Do you want to store the output in a text file? (yes/no): ")

	if store_output.lower() == "yes" or store_output.lower() == "y":
		# Ask for the name of the text file
		file_name = input("Enter the name of the text file: ")

		# Create the text file if it doesn't exist
		with open(file_name, "a") as file:  # Use "a" mode to append to the file
			if output_format.lower() == "yes" or output_format.lower() == "y":
				file.write(tabulate(table, headers=field_names))
			else:
				for row in rows_with_value:
					file.write(str(row) + "\n")


elif selected_option == 3:
	# Retrieve all values from the sheet
	all_values = []
	for row in sheet.iter_rows(min_row=2):
		for cell in row:
			all_values.append(cell.value)

	# Print all the values in a table format
	table = []
	for i in range(0, len(all_values), len(field_names)):
		table.append(all_values[i:i+len(field_names)])
	print(tabulate(table, headers=field_names))

	# Ask the user if they want to store the output in a text file
	store_output = input("Do you want to store the output in a text file? (yes/no): ")

	if store_output.lower() == "yes" or store_output.lower() == "y":
		# Ask for the name of the text file
		file_name = input("Enter the name of the text file: ")

		# Create the text file if it doesn't exist
		with open(file_name, "w") as file:
			file.write(tabulate(table, headers=field_names))

# Close the workbook
workbook.close()
